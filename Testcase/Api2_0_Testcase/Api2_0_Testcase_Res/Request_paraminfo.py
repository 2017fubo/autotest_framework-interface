# -*- coding:utf-8 -*-
__author__ = '付波'
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter
from rizhi import log_config as a
import logging

class ReadWriter_ecxel:

    def __init__(self):
        self.casefile_path  = "E:\SinoBBD\AUTOTEST\Web_Autoest_FW\Testcase\Api2_0_Testcase\API2.0业务测试用例_ver1122.xlsx"
        logging.info(self.casefile_path)
        self.outsheet = '接口测试用例_Out'
        self.intlsheet = '接口测试用例_internal'

    def read_caseparam_int(self,TestName):
        ##打开Testcase excel
        try:
            logging.info("打开excel")
            wb = load_workbook(self.casefile_path)
            self.ws_sheet = wb.get_sheet_by_name(self.intlsheet)
        except Exception as e:
            logging.info("打开excel error%s" %e)
        #获取接口请求参数并返回
        try:
            row_sum = len(self.ws_sheet.rows)###总行数
            #columns_sum = len(ws_sheet.columns)
            expect_data = {}
            #print(row_sum,columns_sum)
            ##匹配用例中的Test urlName值（H列），并获取对应的请求参数（J列）

            logging.info("匹配用例中的TestName值，并获取对应的请求参数")
            for row in range (1,row_sum):
                request_name_temp = self.ws_sheet.cell('H%s' %row).value
                logging.info(request_name_temp)
                if request_name_temp == TestName:
                    logging.info('匹配成功，获取request_param...')
                    #请求request_param
                    self.request_param = self.ws_sheet.cell('J%s' %row).value#请求参数
                    self.request_sys = self.ws_sheet.cell('C%s' %row).value#请求系统
                    self.http_method = self.ws_sheet.cell('G%s' %row).value#请求方法
                    break
                else:
                    row += 1
        except Exception as e:
            logging.info('读取excel信息错误 %s' %e)
        return self.request_param,self.request_sys,self.http_method

    def read_caseparam_out(self,TestName):
        try:
            logging.info("打开excel")
            wb = load_workbook(self.casefile_path)
            self.ws_sheet = wb.get_sheet_by_name(self.outsheet)
        except Exception as e:
            logging.info("打开excel error%s" % e)
        # 获取接口请求参数并返回
        try:
            row_sum = len(self.ws_sheet.rows)  ###总行数
            # columns_sum = len(ws_sheet.columns)
            expect_data = {}
            logging.info("匹配用例中的TestName值，并获取对应的请求参数")
            for row in range(1, row_sum):
                request_name_temp = self.ws_sheet.cell('H%s' % row).value
                logging.info(request_name_temp)
                if request_name_temp == TestName:
                    logging.info('匹配成功，获取request_param...')
                    # 请求request_param
                    self.request_param = self.ws_sheet.cell('J%s' % row).value  # 请求参数
                    self.http_method = self.ws_sheet.cell('G%s' % row).value  # 请求方法
                    break
                else:
                    row += 1
        except Exception as e:
            logging.info('读取excel信息错误 %s' % e)
        # logging.info(self.request_param,self.http_method)

        return self.request_param,self.http_method
    # def read_excel_caseinfo(self,TestName):
    #     ##打开excel
    #     try:
    #         logging.info("打开excel")
    #         wb = load_workbook(self.casefile_path)
    #         ws_sheet = wb.get_sheet_by_name(self.testsheetname)
    #     except Exception as e:
    #         logging.info("打开excel error%s" %e)
    #
    #     # case_list = []
    #     #获取接口数量
    #     try:
    #         row_sum = len(ws_sheet.rows)
    #         columns_sum = len(ws_sheet.columns)
    #         expect_data = {}
    #         #print(row_sum,columns_sum)
    #         ##匹配用例中的TestName值，并获取对应的各项值
    #         logging.info("匹配用例中的TestName值，并获取对应的各项值")
    #         for row in range (2,row_sum):
    #             testname_temp = ws_sheet.cell('C%s' %row).value
    #             print(testname_temp)
    #             if testname_temp == TestName:
    #                 logging.info('匹配成功，获取接口相关信息。')
    #                 #请求url
    #                 request_url = ws_sheet.cell('G%s' %row).value#线上接口地址
    #                 request_param = ws_sheet.cell('I%s' %row).value#请求参数
    #                 expect_code = ws_sheet.cell('K%s' %row).value#预期code
    #                 expect_data = ws_sheet.cell('J%s' %row).value#返回data key and val 值
    #                 expect_msg = ws_sheet.cell('L%s' %row).value#预期message
    #                 logging.info("返回data key and val 值:%s" %expect_data)
    #             else:
    #                 row += 1
    #     except Exception as e:
    #         logging.info('读取excel信息错误 %s' %e)
    #
    #     return request_url,request_param,expect_code,expect_data,expect_msg
    #
    # def writer_excel_testresult(self,case_list,result_info_list):
    #     log_msg = a.RiZhi()
    #     log_msg.log_def()
    #     try:
    #         wb = load_workbook(self.casefile_path)
    #         ew = ExcelWriter(workbook=wb)
    #         ws_sheet = wb.get_sheet_by_name(self.testsheetname)
    #         row_sum = len(ws_sheet.rows)
    #         #tmp_case_list = self.read_excel_caseinfo()
    #         #print('excel中读取的caselist',tmp_case_list)
    #         print('执行caselist:',case_list)
    #         print('测试结果：',result_info_list)
    #         print('开始回填结果判断')
    #         i = 0
    #         for case_id in case_list:
    #             #for case_id_tmp in tmp_case_list:
    #                 #if case_id_tmp == case_id:
    #             for row in range (3,row_sum):
    #                 case_id_TM = ws_sheet.cell('D%s' % row).value
    #                 if case_id ==case_id_TM:
    #                     result_info_tmp = result_info_list[i]
    #                     ws_sheet.cell('o%s' % row).value = result_info_tmp
    #                     print('%s的结果回填完毕'%case_id)
    #                     i = i+1
    #                     print(i)
    #                 else:
    #                     print('该case未执行，继续往下填写测试结果')
    #                 #else:
    #                     #print('执行case与用例case不匹配，继续查找。。。')
    #         ew.save('test_result' + '.xlsx')
    #     except Exception as e:
    #         print("读取error %s" % e)
    #         #logger.info("%s" % e)
    #
    #
    # def read_excel(self, row, col):
    #     try:
    #         logging.info('获取登录信息')
    #         wb = load_workbook(self.casefile_path)
    #         logging.info(wb)
    #         ws_sheet = wb.get_sheet_by_name(self.testsheetname)
    #         logging.info(ws_sheet)
    #         login_username = ws_sheet.cell('%s%s' % (col, row)).value
    #         logging.info(login_username)
    #         row += 1
    #         login_password = ws_sheet.cell('%s%s' % (col, row)).value
    #         logging.info(login_password)
    #
    #     except Exception as e:
    #         logging.info("读取error %s" % e)
    #         # logger.info("%s" % e)
    #
    #     return login_username, login_password